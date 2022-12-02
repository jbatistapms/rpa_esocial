import csv, decimal, os, sys, unicodedata, uuid
from collections import defaultdict, OrderedDict
from datetime import datetime
from hashlib import md5
from pathlib import Path, WindowsPath
from typing import Dict, Final, List, Optional, Tuple, Union

import xlrd
from dotenv import load_dotenv
from loguru import logger
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from tinydb import where

from core import bd, destino, origem

load_dotenv()

COMPETENCIA = os.getenv('COMPETENCIA') # mm/yyyy

ctx_decimal = decimal.getcontext()
ctx_decimal.rounding = decimal.ROUND_HALF_EVEN


class Arquivo(object):
    def __init__(self, loc: WindowsPath) -> None:
        self.__loc = loc
        self.__logs_pln = []
        self.__recibos: List[Recibo] = []
        self.id = None
        if loc.suffix == '.txt':
            try:
                self.id = str(uuid.UUID(loc.stem))
                with loc.open() as loc:
                    for ln in loc.readlines():
                        if ln.startswith("tipo»R"):
                            self.__recibos.append(Recibo.novo_de_texto(text=ln))
            except:
                self.id = None
                logger.error(f"Arquivo '{loc.stem}' com nome inválido.")
        elif loc.suffix == '.xls':
            recibo = Recibo.novo_de_planilha(loc=loc)
            if recibo:
                self.id = str(uuid.UUID(md5(loc.read_bytes()).hexdigest()))
                self.__recibos.append(recibo)
        elif '.fmss' in loc.suffixes and loc.suffix == '.csv':
            ano, mes, _ = loc.stem.split('.')
            with loc.open() as arq:
                csv_arq = csv.DictReader(arq, delimiter=';')
                for linha in csv_arq:
                    self.__recibos.append(Recibo.novo_de_csv(
                        dados=linha,
                        dt_base=f"01/{mes}/{ano}"
                    ))
                    self.id = str(uuid.UUID(md5(loc.read_bytes()).hexdigest()))
    
    def __repr__(self) -> str:
        return repr(self.__loc)
    
    def __str__(self) -> str:
        return str(self.__loc)
    
    def extensao(self) -> str:
        return self.__loc.suffix
    
    def logs_planilha(self) -> List:
        return self.__logs_pln
    
    def recibos(self) -> List['Recibo']:
        return self.__recibos
    
    def salvar(self) -> None:
        for rec in self.__recibos:
            rec.salvar()
        bd.arquivos.insert({'id': self.id})
    

class Coletor(object):
    def __init__(self, origem: WindowsPath) -> None:
        self.__origem = origem
        self.__origem.mkdir(exist_ok=True)
        self.__dir_recibos = self.__origem.joinpath("Recibos")
        self.__dir_recibos.mkdir(exist_ok=True)
        self.__dir_qcadastral = self.__origem.joinpath("Qualificação cadastral")
        self.__dir_qcadastral.mkdir(exist_ok=True)
        self.qcadastral = QualificacaoCadastral(dir_ori=self.__dir_qcadastral, pln=None)
    
    def arquivos(self, dir: WindowsPath) -> List[Arquivo]:
        arquivos = []
        for item in dir.iterdir():
            if item.is_dir():
                arquivos += self.arquivos(dir=item)
            else:
                arquivo = Arquivo(loc=item)
                if arquivo.id and bd.arquivos.get(where('id')==arquivo.id) is None:
                    arquivos.append(arquivo)
        return arquivos
    
    def salvar(self) -> None:
        for arq in self.arquivos(dir=self.__dir_recibos):
            arq.salvar()
        self.qcadastral.importar()


class Controlador(object):
    def __init__(self, destino: WindowsPath) -> None:
        self.__destino: WindowsPath = destino
        self.__destino.mkdir(exist_ok=True)
        mes, ano = COMPETENCIA.split('/')
        self.__loc_recibos = self.__destino.joinpath(f'{ano}.{mes}.xlsx')
        self.pln_pessoas = PlanilhaPessoas(loc=self.__destino.joinpath('Pessoas.xlsx'))
        self.pln_recibos = PlanilhaRecibos(loc=self.__loc_recibos)
        self.qcadastral = QualificacaoCadastral(dir_ori=None, pln=self.pln_pessoas)
    
    def exportar_pessoas(self) -> None:
        # Exportando novas pessoas
        condicao_bd = (
            (where('comp_inicial') == COMPETENCIA) &
            (where('exportado') == False)
        )
        for dados in bd.pessoas.search(condicao_bd):
            self.pln_pessoas.inserir(registro=dados)
        self.pln_pessoas.gravar()
        bd.pessoas.update({'exportado': True}, cond=condicao_bd)
        # Exportando atualização de pessoas
        condicao_bd = (
            (where('atualizado') == True) &
            (where('exportado') == True)
        )
        for dados in bd.pessoas.search(condicao_bd):
            self.pln_pessoas.atualizar(dados['cpf'], dados=dados)
        self.pln_pessoas.gravar()
        bd.pessoas.update({'atualizado': False}, cond=condicao_bd)
        self.qcadastral.exportar()
    
    def exportar_recibos(self) -> None:
        condicao_bd = (
            (where('dt_final').search(f'{COMPETENCIA}')) &
            (where('exportado') == False)
        )
        for dados in bd.recibos.search(condicao_bd):
            self.pln_recibos.inserir(registro=dados)
        # Atualizar dados de pessoas
        for reg in self.pln_pessoas.registros():
            self.pln_recibos.atualizar_cpf(cpf=reg['cpf'], dados=reg)
        self.pln_recibos.gravar()
        bd.recibos.update({'exportado': True}, cond=condicao_bd)


class Recibo(dict):
    def __init__(self, **dados) -> None:
        super().__init__(dados)
        self.__erros = []
        self.__tratar_dados()
        self.__dados_pessoa = copiar_dict(
            obj=self,
            inc=['cpf', 'dt_nascimento', 'nis_pis', 'nome', 'qcadastral', 'exportado']
        )
        self.__dados_pessoa['comp_inicial'] = self['dt_inicial'][3:]
    
    def __limpar_doc(self, cmp: str, cont: int) -> None:
        valor = self[cmp]
        cmp_exc = PlanilhaRecibos.conversor_ao_contrario[cmp]
        if isinstance(valor, str):
            vl_retorno = self[cmp]
            for caracter in '.-/_':
                vl_retorno = vl_retorno.replace(caracter, '')
            if vl_retorno.isdigit() and len(vl_retorno) == cont:
                valor = int(vl_retorno)
            else:
                self.__erros.append(f"Número de '{cmp_exc}' deve possuir {cont} dígitos: {valor}")
        elif isinstance(valor, (int, float)):
            if not valor < 10**cont:
                self.__erros.append(f"Número de '{cmp_exc}' deve possuir {cont} dígitos: {valor}")
        else:
            self.__erros.append(f"Número de '{cmp_exc}' não reconhecido: {valor}")
        self[cmp] = valor

    def __tratar_dados(self) -> dict:
        self['erros'] = ''
        self.__erros = []
        self.update({
            'nome': unicodedata.normalize(
                "NFKD", self['nome']
                ).encode(
                    'ASCII', 'ignore'
                ).decode().upper().strip(),
        })
        # Campos com valores
        for cmp in [
                'dsc_inss', 'dsc_irrf', 'dsc_iss', 'valor', 'v_liquido', 
                'outrosd_valor1', 'outrosd_valor2'
            ]:
            if self[cmp] == '':
                self[cmp] = 0
            valor, erro = normalizar_valor(valor=self[cmp])
            if erro:
                self.__erros.append(erro)
            else:
                self[cmp] = valor
        # Campo com datas
        for cmp in ['dt_final', 'dt_inicial', 'dt_nascimento']:
            valor = normalizar_data(data=self[cmp])
            if valor is None:
                self.__erros.append(
                    f"{PlanilhaRecibos.conversor_ao_contrario[cmp]} "
                    f"ilegível: '{self[cmp]}'"
                )
                self.update({cmp: ''})
            else:
                self.update({cmp: valor})
        # Noemalizar documentos
        self.__limpar_doc(cmp='nis_pis', cont=11)
        self.__limpar_doc(cmp='cpf', cont=11)
        self.__limpar_doc(cmp='cbo', cont=6)
        for erro in self.__erros:
            self['erros'] += erro + ';'
        # Diferença INSS
        valor = decimal.Decimal(self['valor'])
        inss = valor * decimal.Decimal('0.11')
        inss = float(inss.quantize(decimal.Decimal('1.00')))
        self.update({'dif_inss': self['dsc_inss'] - inss})
        # Valor líquido
        v_liquido = self['valor']
        for cmp in ['dsc_inss', 'dsc_irrf', 'dsc_iss', 'outrosd_valor1', 'outrosd_valor2']:
            v_liquido -= self[cmp]
        self.update({'v_liquido': v_liquido})
        # Outros descontos
        if self['outrosd_desc1'] in ['0', 0]:
            self['outrosd_desc1'] = ''
        if self['outrosd_desc2'] in ['0', 0]:
            self['outrosd_desc2'] = ''
    
    def atualizar(self, dados: dict) -> None:
        self.update(dados)
    
    def dados_planilha(self) -> dict:
        rec_temp = OrderedDict()
        for exc, txt in PlanilhaRecibos.conversor.items():
            rec_temp[exc] = self.get(txt, '')
        return list(rec_temp.values())
    
    def erros(self) -> str:
        erros = ''
        for erro in self.__erros:
            erros += erro + ';'
        return erros if len(erros) > 0 else None
    
    @classmethod
    def __novo_dados_padrao(cls) -> dict:
        return {'qcadastral': 'NÃO', 'erros': '', 'exportado': False}
    
    @classmethod
    def novo_de_csv(cls, dados: dict, dt_base: str) -> 'Recibo':
        registro = cls.__novo_dados_padrao()
        registro['nome'] = dados['Prestador de serviço']
        registro['cpf'] = dados['CNPJ/CPF']
        registro['dt_nascimento'] = dados['D. Nasc.']
        registro['nis_pis'] = dados['NIS']
        registro['cbo'] = dados['CBO']
        registro['servico_prestado'] = ''
        registro['dt_inicial'] = dt_base
        registro['dt_final'] = dt_base
        registro['valor'] = dados['BRUTO']
        registro['dsc_inss'] = dados['INSS']
        registro['dsc_irrf'] = dados['IRRF']
        registro['dsc_iss'] = dados['ISS']
        registro['outrosd_desc1'] = '0'
        registro['outrosd_valor1'] = '0'
        registro['outrosd_desc2'] = '0'
        registro['outrosd_valor2'] = '0'
        registro['v_liquido'] = dados['Líquido']
        registro['orgao'] = 'Fundo Municipal de Saúde de Sapucaia'
        return cls(**registro)
    
    @classmethod
    def novo_de_planilha(cls, loc: WindowsPath) -> Optional['Recibo']:
        book: xlrd.Book = xlrd.open_workbook(loc)
        sheet: xlrd.sheet.Sheet = book.sheet_by_name('RPA')
        if sheet.cell_value(rowx=60, colx=0) != 'Responsavel/Tomador do Serviço':
            logger.error(f"Planilha em '{loc}' não está formatada corretamente.")
            return None
        dt_inicial = xlrd.xldate.xldate_as_tuple(sheet.cell_value(rowx=30, colx=1), 0)
        dt_final = xlrd.xldate.xldate_as_tuple(sheet.cell_value(rowx=30, colx=3), 0)
        dados = cls.__novo_dados_padrao()
        dados['nome'] = sheet.cell_value(rowx=13, colx=1)
        dados['cpf'] = sheet.cell_value(rowx=16, colx=1)
        dados['dt_nascimento'] = ''
        dados['nis_pis'] = sheet.cell_value(rowx=17, colx=1)
        dados['cbo'] = sheet.cell_value(rowx=28, colx=1)
        dados['servico_prestado'] = sheet.cell_value(rowx=27, colx=1)
        dados['dt_inicial'] = '{2:>02}/{1:>02}/{0}'.format(*dt_inicial)
        dados['dt_final'] = '{2:>02}/{1:>02}/{0}'.format(*dt_final)
        dados['valor'] = sheet.cell_value(rowx=29, colx=1)
        dados['dsc_inss'] = sheet.cell_value(rowx=34, colx=1)
        dados['dsc_irrf'] = sheet.cell_value(rowx=38, colx=1)
        dados['dsc_iss'] = sheet.cell_value(rowx=40, colx=1)
        dados['outrosd_desc1'] = sheet.cell_value(rowx=42, colx=0)
        dados['outrosd_valor1'] = sheet.cell_value(rowx=42, colx=1)
        dados['outrosd_desc2'] = sheet.cell_value(rowx=43, colx=0)
        dados['outrosd_valor2'] = sheet.cell_value(rowx=43, colx=1)
        dados['v_liquido'] = sheet.cell_value(rowx=44, colx=1)
        dados['orgao'] = sheet.cell_value(rowx=7, colx=1)
        return cls(**dados)
    
    @classmethod
    def novo_de_texto(cls, text: str) -> 'Recibo':
        dados = cls.__novo_dados_padrao()
        for ch_vl in text.split('«')[:-1]:
            ch, vl = ch_vl.split('»')
            dados.update({ch: vl})
        return cls(**dados)
    
    def salvar(self) -> None:
        pessoa_db = bd.pessoas.get(where('cpf')==self['cpf'])
        if pessoa_db is None:
            bd.pessoas.insert(self.__dados_pessoa)
        assinatura_recibo = (str(self['cbo']) +
            str(self['cpf']) +
            str(self['dt_final']) +
            str(self['orgao']) +
            str(self['valor'])
        )
        self['id'] = str(uuid.UUID(md5(assinatura_recibo.encode()).hexdigest()))
        recibo_db = bd.recibos.get(where('id')==self['id'])
        if recibo_db is None:
            bd.recibos.insert(self)


class ObjetoPlanilha(dict):
    def atualizar(self, dados: dict) -> None:
        self.update(dados)


class Planilha(object):
    classe = ObjetoPlanilha
    conversor: dict
    estilos: dict
    identificador: str
    titulo: str

    def __init__(self, loc: WindowsPath) -> None:
        self.loc: Final = loc
        self.__registros: List[ObjetoPlanilha]
        if loc.exists():
            self.__ler()
        else:
            self.__registros = defaultdict(dict)
    
    def __ler(self) -> None:
        wb = load_workbook(self.loc)
        ws = wb[self.titulo]
        lista_linhas = []
        for row in ws.rows:
            linha = []
            for cell in row:
                linha.append(cell.value)
            lista_linhas.append(linha)
        cabecalho = lista_linhas.pop(0)
        self.__registros = defaultdict()
        for linha in lista_linhas:
            dados = defaultdict()
            for exc, txt in self.conversor.items():
                try:
                    idx = cabecalho.index(exc)
                except:
                    logger.error(f"Coluna {exc} inexistente em {self.loc}")
                    sys.exit()
                dados[txt] = linha[idx]
            if dados[self.identificador]:
                self.__registros[dados[self.identificador]] = self.classe(**dados)
    
    def atualizar(self, chave: str, dados: dict) -> None:
        if chave in self.__registros:
            self.__registros[chave].atualizar(dados)
    
    def dados_gravar(self, reg: dict) -> list:
        return [reg[col] for col in self.conversor.values()]
    
    def gravar(self) -> None:
        wb = Workbook()
        ws = wb.active
        ws.title = self.titulo

        # Escrever cabeçalho
        cbc = list(self.conversor.keys())
        ws.append(cbc)
        
        # Escrever linhas
        linhas = self.registros()
        for reg in linhas:
            ws.append(self.dados_gravar(reg=reg))
        
        colunas = list(ws.columns)
        
        # Criar tabela e definir estilo
        style = TableStyleInfo(
            name="TableStyleLight15",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        column = list(ws.columns)[-1][0].column_letter
        tab = Table(
            displayName=self.titulo,
            ref=f"A1:{column}{len(self.__registros)+1}",
            tableStyleInfo=style,
        )
        ws.add_table(tab)

        # Centralizar cabeçalho
        for row in ws.iter_rows(min_row=1, max_col=len(cbc), max_row=1):
            for cell in row:
                cell.alignment = Alignment(horizontal='center')

        # Redimensionamento
        for col_n, column_cells in enumerate(colunas):
            new_column_length = max(
                [len(str(cell.value)) for cell in column_cells] +
                [len(cbc[col_n])+5]
            )
            new_column_letter = get_column_letter(column_cells[0].column)
            if new_column_length > 0:
                ws.column_dimensions[new_column_letter].width = new_column_length*1.25
        
        # Aplicar estilos
        for stl in self.tipo_estilos():
            wb.add_named_style(stl)
        for stl, colunas in self.estilos.items():
            for col_nome in colunas:
                idx = cbc.index(col_nome) + 1
                for col in ws.iter_cols(min_col=idx, max_col=idx, min_row=2, max_row=len(linhas)*2):
                    for cell in col:
                        cell.style = stl

        # Salvar arquivo
        wb.save(self.loc)
    
    def inserir(self, registro: dict) -> None:
        if not registro[self.identificador] in self.__registros:
            self.__registros[registro[self.identificador]] = self.classe(**registro)
    
    def registros(self) -> List:
        return list(self.__registros.values())
    
    def tipo_estilos(self) -> dict:
        alinhamento_centro = Alignment(horizontal='center')
        cbo = NamedStyle(
            name='cbo',
            number_format='0000"-"00',
            alignment=alinhamento_centro,
        )
        cpf = NamedStyle(
            name='cpf',
            number_format='000"."000"."000"-"00',
            alignment=alinhamento_centro,
        )
        data = NamedStyle(
            name='data',
            number_format='dd/mm/yyyy',
            alignment=alinhamento_centro,
        )
        moeda = NamedStyle(
            name='moeda',
            number_format='#,##0.00',
        )
        nit = NamedStyle(
            name='nit',
            number_format='000"."00000"."00"-"0',
            alignment=alinhamento_centro,
        )
        centro = NamedStyle(
            name='centro',
            alignment=alinhamento_centro,
        )
        return [cbo, cpf, data, moeda, nit, centro]


class PlanilhaPessoas(Planilha):
    conversor = {
        'Nome': 'nome',
        'CPF': 'cpf',
        'NIS/PIS': 'nis_pis',
        'Data de nascimento': 'dt_nascimento',
        'Competência Inicial': 'comp_inicial',
        'QCadastral': 'qcadastral',
    }
    estilos = {
        'cpf': ['CPF'],
        'data': ['Data de nascimento'],
        'nit': ['NIS/PIS'],
        'centro': ['QCadastral', 'Competência Inicial']
    }
    identificador = 'cpf'
    titulo = 'Pessoas'


class PlanilhaRecibos(Planilha):
    classe = Recibo
    conversor = {
        'Identificador': 'id',
        'Nome': 'nome',
        'CPF': 'cpf',
        'Data de nascimento': 'dt_nascimento',
        'NIS/PIS': 'nis_pis',
        'CBO': 'cbo',
        'Serviço prestado': 'servico_prestado',
        'Data inicial': 'dt_inicial',
        'Data final': 'dt_final',
        'Valor': 'valor',
        'Dsc. INSS': 'dsc_inss',
        'Dif. INSS': 'dif_inss',
        'Dsc. IRRF': 'dsc_irrf',
        'Dsc. ISS': 'dsc_iss',
        'Descrição - Outros descontos I': 'outrosd_desc1',
        'Valor - Outros descontos I': 'outrosd_valor1',
        'Descrição - Outros descontos II': 'outrosd_desc2',
        'Valor - Outros descontos II': 'outrosd_valor2',
        'Líquido': 'v_liquido',
        'Órgão contratante': 'orgao',
        'Data do pagamento': 'dt_pgto',
        'Erros': 'erros',
        'Enviar': 'enviar',
        'Demonstrativo': 'demonstrativo',
        'ID S-1200': 'id_s1200',
        'Protocolo de envio S-1200': 'protocolo_s1200',
        'Recibo de envio S-1200': 'recibo_s1200',
        'Erro no envio S-1200': 'erro_s1200',
        'ID S-1210': 'id_s1210',
        'Protocolo de envio S-1210': 'protocolo_s1210',
        'Recibo de envio S-1210': 'recibo_s1210',
        'Erro no envio S-1210': 'erro_s1210',
    }
    conversor_ao_contrario = {v: k for k, v in conversor.items()}
    estilos = {
        'data': ['Data inicial', 'Data de nascimento', 'Data final'],
        'cpf': ['CPF'],
        'nit': ['NIS/PIS'],
        'cbo': ['CBO'],
        'moeda': [
            'Valor', 'Dsc. INSS', 'Dif. INSS', 'Dsc. IRRF', 'Dsc. ISS', 
            'Valor - Outros descontos I', 'Valor - Outros descontos II',
            'Líquido'
        ],
        'centro': [
            'Enviar', 'ID S-1200', 'Protocolo de envio S-1200', 'Recibo de envio S-1200',
            'ID S-1210', 'Protocolo de envio S-1210', 'Recibo de envio S-1210',
            'Identificador', 'Demonstrativo', 'Data do pagamento',
        ]
    }
    identificador = 'id'
    titulo = 'Recibos'

    def __init__(self, loc: WindowsPath) -> None:
        super().__init__(loc)
        self.__indice_cpf = defaultdict(list)
        for reg in self.registros():
            self.__indice_cpf[reg['cpf']].append(reg['id'])

    def atualizar_cpf(self, cpf: str, dados: dict) -> None:
        for id_ in self.__indice_cpf.get(cpf, []):
            self.atualizar(id_, dados)
    
    def dados_gravar(self, reg: Recibo) -> dict:
        return reg.dados_planilha()


class QualificacaoCadastral(object):
    def __init__(self, dir_ori: WindowsPath, pln: PlanilhaPessoas) -> None:
        self.dir_ori = dir_ori
        self.pln_pessoas = pln
    
    def exportar(self) -> None:
        if self.pln_pessoas is None:
            logger.warning("Instância de 'PlanilhaPessoas' não foi fornecida.")
        else:
            with self.pln_pessoas.loc.parent.joinpath('qualificacao_cadastral.txt').open('w') as a:
                for ps in self.pln_pessoas.registros():
                    if ps['qcadastral'] != 'SIM' and ps['comp_inicial'] == COMPETENCIA:
                        if ps['cpf'] and ps['nome'] and ps['dt_nascimento']:
                            a.write(
                                f"{ps['cpf']:>011};13333333332;{ps['nome']};"
                                f"{limpar_data(ps['dt_nascimento'])}\n"
                            )
                        else:
                            logger.warning(
                                f"Cadastro de {ps['nome']} está incompleto "
                                "para qualificação cadastral."
                            )
    
    def importar(self) -> None:
        logger.info("Iniciando importação de retornos de qualificação cadastral.")
        col_erros = [
            'COD_CNIS_CPF',
            'COD_CNIS_CPF_NAO_INF',
            'COD_CNIS_DN',
            'COD_CNIS_NIS',
            'COD_CNIS_OBITO',
            'COD_CPF_CANCELADO',
            'COD_CPF_DN',
            'COD_CPF_INV',
            'COD_CPF_NAO_CONSTA',
            'COD_CPF_NOME',
            'COD_CPF_NULO',
            'COD_CPF_SUSPENSO',
            'COD_DN_INV',
            'COD_NIS_INV',
            'COD_NOME_INV',
        ]
        lst_registros = defaultdict(dict)
        ids_arqs = []
        for item in self.dir_ori.iterdir():
            if item.is_dir() or not item.suffix == '.PROCESSADO':
                continue
            id_arq = str(uuid.UUID(md5(item.read_bytes()).hexdigest()))
            if bd.arquivos.get(where('id')==id_arq) is not None:
                continue
            logger.info(f"Novo arquivo encontrado em {item}.")
            arq = item.open('r', newline='\n')
            arq_csv = csv.DictReader(arq, delimiter=';')
            for reg in arq_csv:
                if reg['NOME'] == None:
                    continue
                erros = defaultdict()
                for cod_erro in col_erros:
                    if reg[cod_erro] != '0':
                        erros[cod_erro] = reg[cod_erro]
                # Correção de nome
                nome = reg['COD_CPF_NOME']
                if nome != '0' and nome.startswith('1 - ') and nome[4:] != '':
                    lst_registros[reg['CPF']]['nome'] = nome[4:]
                    erros.pop('COD_CPF_NOME')
                if erros:
                    lst_registros[reg['CPF']] = {
                        'qcadastral': ';'.join([f'{cod}: {err}' for cod, err in erros.items()]),
                        'exportado': False,
                        'atualizado': True,
                    }
                else:
                    lst_registros[reg['CPF']] = {
                        'qcadastral': 'SIM',
                        'exportado': False,
                        'atualizado': True,
                    }
            ids_arqs.append(id_arq)
        for cpf, reg in lst_registros.items():
            bd.pessoas.update(reg, cond=(where('cpf')==int(cpf)))
        bd.arquivos.insert_multiple([{'id': id_} for id_ in ids_arqs])

def copiar_dict(obj: dict, exc: Optional[List[str]]=None, inc: Optional[List[str]]=None) -> dict:
    rtrn = {}
    if inc:
        for opc in inc:
            if opc in obj:
                rtrn.update({opc: obj[opc]})
    elif exc:
        rtrn = obj.copy()
        for opc in exc:
            if opc in obj:
                rtrn.pop(opc)
    else:
        rtrn = obj.copy()
    return rtrn

def limpar_doc(doc_txt: str) -> str:
    if isinstance(doc_txt, str):
        return doc_txt.replace('.', '').replace('-', '').replace('/', '')
    return ''

def limpar_data(dt: Union[str, datetime]) -> str:
    if isinstance(dt, str):
        return dt.replace('.', '').replace('-', '').replace('/', '')
    elif isinstance(dt, datetime):
        return dt.strftime('%d%m%Y')
    else:
        raise ValueError("Parâmetro 'dt' deve ser do tipo str ou datetime.")

def normalizar_data(data: Union[str,datetime]) -> str:
    if isinstance(data, datetime):
        return data.strftime('%d/%m/%Y')
    elif isinstance(data, str):
        data = data.replace('/', '')
        if len(data) == 8 and data.isdigit():
            return f"{data[0:2]}/{data[2:4]}/{data[4:8]}"
    return None

def normalizar_valor(valor: str) -> Tuple[Optional[float],Optional[str]]:
    if valor in ['', '0', 0]:
        return 0, None
    elif isinstance(valor, (int, float)):
        return valor, None
    elif valor.startswith('R$ '):
        try:
            _, valor = valor.split()
            return float(valor.replace('.', '').replace(',', '.')), None
        except Exception as erro:
            return None, f"Valor '{valor}' não corresponde ao esperado: {erro}"
    else:
        return None, f"Valor '{valor}' não corresponde ao esperado."

def salvar_dados(dados: List[dict], dst: Path, fieldnames: Optional[List]=None) -> None:
    with dst.open('w', newline='') as arq:
        if dados:
            fieldnames = fieldnames or list(dados[0].keys())
            arq_csv = csv.DictWriter(arq, fieldnames=fieldnames, delimiter=';')
            arq_csv.writeheader()
            arq_csv.writerows(dados)
        else:
            arq.write('')

if __name__ == '__main__':
    colt = Coletor(origem=origem)
    colt.salvar()
    exp = Controlador(destino=destino)
    exp.exportar_pessoas()
    exp.exportar_recibos()